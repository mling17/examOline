import uuid, os
from django.shortcuts import render, HttpResponse, redirect
from django.http import JsonResponse
from question_bank import models
from question_bank.forms.bank import BankAddModelForm
from django.urls import reverse
from django.db import transaction
from openpyxl import load_workbook
from question_bank.forms.upload_file import UploadFileModelForm
from utils.pagination import Pagination
from django.conf import settings


# Create your views here.
def bank_list(request):
    if request.method == 'GET':
        banks = models.QuestionBank.objects.filter(is_valid=True).all()
        print(banks)
        all_count = models.QuestionBank.objects.filter(is_valid=True).count()
        print(all_count)
        query_params = request.GET.copy()
        query_params._mutable = True
        pager = Pagination(
            current_page=request.GET.get('page'),
            all_count=all_count,
            base_url=request.path_info,
            query_params=query_params,
            per_page=settings.PER_PAGE_COUNT,
        )
        bank_data_list = banks[pager.start:pager.end]
        return render(request, 'bank_list.html', {'banks': bank_data_list, 'pager': pager})


def bank_add(reqeust):
    if reqeust.method == 'GET':
        form = BankAddModelForm()
        return render(reqeust, 'change.html', {'form': form})
    form = BankAddModelForm(data=reqeust.POST)
    if form.is_valid():
        title = form.cleaned_data.get('title')
        content = form.cleaned_data.get('content')
        cover = form.cleaned_data.get('cover')
        subject_name = form.cleaned_data.get('subject_name')
        models.QuestionBank.objects.create(title=title, content=content, cover=cover, subject_name=subject_name)
        return redirect('question_bank:bank_list')
    return render(reqeust, 'change.html', {'form': form})


def edit(request, pk):
    response = {'status': 'success'}
    bank_obj = models.QuestionBank.objects.filter(id=pk).first()
    if not bank_obj:
        response = {'status': 'error', 'msg': '未找到此条目'}
        return JsonResponse(response)
    if request.method == 'GET':
        response['id'] = bank_obj.id
        response['title'] = bank_obj.title
        response['content'] = bank_obj.content
        response['subject_name'] = bank_obj.subject_name
        return JsonResponse(response)
    title = request.POST.get('title', '')
    content = request.POST.get('content', '')
    subject_name = request.POST.get('subject_name', '')
    if all((title, content, subject_name)):
        response['title'] = title
        response['content'] = content
        response['subject_name'] = subject_name
        bank_obj.title = title
        bank_obj.content = content
        bank_obj.subject_name = subject_name
        bank_obj.save()
        return JsonResponse(response)
    response = {'status': 'error', 'msg': '数据不能为空'}
    return JsonResponse(response)


def delete(request, pk):
    # origin_list_url = self.reverse_list_url(*args, **kwargs)
    origin_list_url = reverse('question_bank:bank_list')
    if request.method == 'GET':
        return render(request, 'bank_delete.html', {'cancel': origin_list_url})
    models.QuestionBank.objects.filter(pk=pk).delete()
    return redirect(origin_list_url)


def load_question(file):
    wb = load_workbook(filename=file)
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[0]]
    max_row = ws.max_row
    max_column = ws.max_column
    all_list = []
    lis = []
    for row in range(1, max_row + 1):
        dic = {}
        for column in range(1, max_column + 1):
            if row == 1:
                lis.append(ws.cell(row, column).value)
                continue
            dic.setdefault(lis[column - 1], ws.cell(row, column).value)
        if row > 1:
            all_list.append(dic)
    headers = ['bank_title', 'bank_content', 'type', 'title', 'analysis', 'sort_num']
    bank_name = None
    for q_dic in all_list:
        if not all(q_dic.values()) and not q_dic['type']:
            continue
        imp_bank_name = q_dic['bank_title'].strip()
        if bank_name != imp_bank_name:
            bank_name = imp_bank_name
            new_bank_obj = models.QuestionBank()
        new_question_obj = models.Question()
        try:
            with transaction.atomic():
                for key in headers:
                    if key.startswith('bank_'):
                        setattr(new_bank_obj, key.lstrip('bank_'), q_dic[key])
                    else:
                        setattr(new_question_obj, key, q_dic[key])
                        new_question_obj.question_bank = new_bank_obj  # 题与题库关联
                new_bank_obj.save()
                new_question_obj.save()
                if q_dic['type'] in (1, 2, 3):
                    answer = q_dic['answer'].upper().strip()
                    options = eval(q_dic['options'])
                    for option in options:
                        title = option.get('title')
                        if title.upper() == 'C' and q_dic['type'] == 3:
                            break
                        option_obj = models.ChoiceOptions()
                        content = option.get('content')
                        if title.upper().strip() == answer:
                            option_obj.is_answer = True
                        option_obj.title = title
                        option_obj.content = content
                        option_obj.question = new_question_obj  # 选项与题关联
                        option_obj.save()
        except Exception as e:
            print(q_dic)
            print(e)
            return False


def import_question(request):
    if request.method == 'GET':
        form = UploadFileModelForm()
        return render(request, 'test_page.html', {'form': form})
    form = UploadFileModelForm(request.POST, request.FILES)
    if form.is_valid():
        raw_file = form.cleaned_data.get("file")
        origin_name = raw_file.name
        ext = raw_file.name.split('.')[-1]
        new_name = '{}.{}'.format(uuid.uuid4().hex, ext)
        raw_file.name = new_name
        new_file_obj = models.ImportFile()
        new_file_obj.file = raw_file
        new_file_obj.name = origin_name
        new_file_obj.user_id = request.user_info.user.id
        new_file_obj.save()
        result = load_question(raw_file.open())
        if result is False:
            content = {'msg': '导入失败，3秒后跳转...', 'next_pg': reverse('question_bank:bank_list')}
        else:
            content = {'msg': '导入成功，3秒后跳转...', 'next_pg': reverse('question_bank:bank_list')}
        return render(request, 'common_msg.html', context=content)
    return render(request, 'test_page.html', {'form': form})
